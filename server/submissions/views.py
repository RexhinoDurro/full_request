# server/submissions/views.py - COMPLETE ADMIN IMPLEMENTATION

import re
import hashlib
import json
import logging
from datetime import timedelta
from io import BytesIO
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.utils import timezone
from django.core.cache import cache
from django.db import transaction
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count
from django.core.exceptions import ValidationError
from django.utils.html import escape

from rest_framework import generics, status
from rest_framework.decorators import api_view, permission_classes, throttle_classes
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.throttling import UserRateThrottle, AnonRateThrottle
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator

from .models import Submission
from .serializers import SecureSubmissionCreateSerializer, SubmissionListSerializer, SubmissionDetailSerializer
from security_monitoring.models import SecurityEvent
from security_monitoring.utils import get_client_ip, log_security_event

logger = logging.getLogger('security_monitoring')

# Previous SQLInjectionDetector and other classes remain the same...
class SQLInjectionDetector:
    """Advanced SQL injection detection and prevention"""
    
    SQL_INJECTION_PATTERNS = [
        r'\bunion\s+(all\s+)?select\b',
        r'\b(and|or)\s+\d+\s*[=<>!]+\s*\d+\b',
        r'\b1\s*=\s*1\b',
        r'(--|#|/\*|\*/)',
        r'\b(select|insert|update|delete|drop|create|alter|exec|execute)\b',
        r'\binformation_schema\b',
    ]
    
    @classmethod
    def detect_sql_injection(cls, text):
        if not text:
            return False, None, 'NONE'
        
        text = str(text).lower()
        for pattern in cls.SQL_INJECTION_PATTERNS:
            if re.search(pattern, text, re.IGNORECASE):
                if any(dangerous in pattern for dangerous in ['drop', 'delete', 'insert', 'update', 'exec']):
                    severity = 'CRITICAL'
                elif any(medium in pattern for medium in ['union', 'select', 'information_schema']):
                    severity = 'HIGH'
                else:
                    severity = 'MEDIUM'
                return True, pattern, severity
        return False, None, 'NONE'

class SQLSecureRateLimiter:
    @staticmethod
    def is_rate_limited(key, limit, window_seconds):
        try:
            safe_key = re.sub(r'[^a-zA-Z0-9_\-:]', '_', str(key))
            safe_key = f"rate_limit:{safe_key}"
            
            current_time = timezone.now()
            attempts = cache.get(safe_key, [])
            
            cutoff_time = current_time - timedelta(seconds=window_seconds)
            recent_attempts = [
                attempt for attempt in attempts 
                if attempt > cutoff_time.timestamp()
            ]
            
            if len(recent_attempts) >= limit:
                return True
            
            recent_attempts.append(current_time.timestamp())
            cache.set(safe_key, recent_attempts, window_seconds)
            return False
            
        except Exception as e:
            logger.error(f"Rate limiter error: {e}")
            return False

# PUBLIC FORM SUBMISSION ENDPOINT
@api_view(['POST'])
@permission_classes([AllowAny])
@csrf_exempt
@never_cache
def submit_form(request):
    """Public form submission endpoint"""
    
    client_ip = get_client_ip(request)
    
    # SQL injection detection
    try:
        request_body = request.body.decode('utf-8') if request.body else ''
        all_input_text = f"{request_body} {request.META.get('HTTP_USER_AGENT', '')} {request.path}"
        
        is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(all_input_text)
        
        if is_injection:
            log_security_event(
                'SQL_INJECTION_ATTEMPT', 'CRITICAL', 
                client_ip, request.META.get('HTTP_USER_AGENT', ''),
                request.user if request.user.is_authenticated else None,
                f'SQL injection detected: {pattern}'
            )
            return Response({
                'success': False,
                'message': 'Security violation detected. Request blocked.'
            }, status=status.HTTP_403_FORBIDDEN)
    
    except Exception as e:
        logger.error(f"SQL injection detection error: {e}")
        return Response({
            'success': False,
            'message': 'Security check failed. Request blocked.'
        }, status=status.HTTP_403_FORBIDDEN)
    
    # Rate limiting
    safe_ip = re.sub(r'[^0-9\.\:]', '_', client_ip)
    if SQLSecureRateLimiter.is_rate_limited(f"form_submit_{safe_ip}", 3, 3600):
        log_security_event(
            'RATE_LIMIT', 'HIGH',
            client_ip, request.META.get('HTTP_USER_AGENT', ''),
            None, 'Form submission rate limit exceeded'
        )
        return Response({
            'success': False,
            'message': 'Too many requests. Please wait before trying again.'
        }, status=status.HTTP_429_TOO_MANY_REQUESTS)
    
    # Validate form data
    for field, value in request.data.items():
        if value:
            is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(str(value))
            if is_injection:
                log_security_event(
                    'SQL_INJECTION_ATTEMPT', 'CRITICAL',
                    client_ip, request.META.get('HTTP_USER_AGENT', ''),
                    None, f'SQL injection in field {field}: {pattern}'
                )
                return Response({
                    'success': False,
                    'message': 'Invalid data detected. Please check your input.'
                }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        serializer = SecureSubmissionCreateSerializer(data=request.data)
        
        if serializer.is_valid():
            with transaction.atomic():
                submission = serializer.save()
                
                log_security_event(
                    'FORM_SUBMISSION', 'LOW',
                    client_ip, request.META.get('HTTP_USER_AGENT', ''),
                    None, 'Form submitted successfully',
                    {
                        'submission_uuid': str(submission.uuid),
                        'country': submission.country,
                    }
                )
            
            return Response({
                'success': True,
                'message': 'Thank you for your submission. Our team will review your application and contact you within 2-3 business days.',
                'submission_id': str(submission.uuid)
            }, status=status.HTTP_201_CREATED)
        
        else:
            log_security_event(
                'FORM_VALIDATION_ERROR', 'LOW',
                client_ip, request.META.get('HTTP_USER_AGENT', ''),
                None, 'Form validation failed'
            )
            
            return Response({
                'success': False,
                'message': 'Please correct the errors in your form.',
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        logger.error(f"Form submission error: {e}", exc_info=True)
        log_security_event(
            'API_ERROR', 'HIGH',
            client_ip, request.META.get('HTTP_USER_AGENT', ''),
            None, 'Form submission processing error'
        )
        
        return Response({
            'success': False,
            'message': 'Unable to process submission. Please try again later.'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ADMIN ENDPOINTS

@api_view(['GET'])
@permission_classes([IsAdminUser])
@csrf_exempt
@never_cache
def admin_submissions(request):
    """Get paginated list of submissions with filtering"""
    
    try:
        # SQL injection detection on all query parameters
        for param_name, param_value in request.GET.items():
            if param_value:
                is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(str(param_value))
                if is_injection:
                    log_security_event(
                        'SQL_INJECTION_ATTEMPT', 'CRITICAL',
                        get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                        request.user, f'SQL injection in parameter {param_name}: {pattern}'
                    )
                    return Response({
                        'success': False,
                        'message': 'Security violation detected in query parameters'
                    }, status=status.HTTP_403_FORBIDDEN)
        
        # Whitelist approach for filters
        ALLOWED_FILTERS = {
            'search': r'^[a-zA-Z0-9\s@\.\-\_]{1,100}$',
            'date_from': r'^\d{4}-\d{2}-\d{2}$',
            'date_to': r'^\d{4}-\d{2}-\d{2}$',
            'country': r'^[A-Z]{2}$',
            'service_type': r'^[a-zA-Z\s\-]{1,50}$',
            'issue_timeframe': r'^[a-zA-Z\s\-]{1,50}$',
            'primary_goal': r'^[a-zA-Z\s\-]{1,50}$',
            'communication_method': r'^[a-zA-Z\s\-]{1,50}$',
        }
        
        validated_filters = {}
        for filter_key, pattern in ALLOWED_FILTERS.items():
            if filter_key in request.GET:
                filter_value = str(request.GET[filter_key]).strip()
                if re.match(pattern, filter_value):
                    validated_filters[filter_key] = filter_value
        
        # Start with base queryset
        queryset = Submission.objects.all()
        
        # Apply validated filters using Django ORM
        if validated_filters.get('search'):
            search_term = validated_filters['search']
            queryset = queryset.filter(
                Q(name__icontains=search_term) |
                Q(email__icontains=search_term) |
                Q(step1__icontains=search_term) |
                Q(step8__icontains=search_term)
            )
        
        if validated_filters.get('date_from'):
            try:
                start_date = timezone.datetime.strptime(validated_filters['date_from'], '%Y-%m-%d').date()
                queryset = queryset.filter(submitted_at__date__gte=start_date)
            except ValueError:
                pass
        
        if validated_filters.get('date_to'):
            try:
                end_date = timezone.datetime.strptime(validated_filters['date_to'], '%Y-%m-%d').date()
                queryset = queryset.filter(submitted_at__date__lte=end_date)
            except ValueError:
                pass
        
        if validated_filters.get('country'):
            queryset = queryset.filter(country=validated_filters['country'])
        
        if validated_filters.get('service_type'):
            queryset = queryset.filter(step2__icontains=validated_filters['service_type'])
        
        if validated_filters.get('issue_timeframe'):
            queryset = queryset.filter(step3__icontains=validated_filters['issue_timeframe'])
        
        # Pagination
        page = int(request.GET.get('page', 1))
        page_size = min(int(request.GET.get('page_size', 20)), 100)  # Max 100 per page
        
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        
        total_count = queryset.count()
        submissions = queryset.order_by('-submitted_at')[start_index:end_index]
        
        # Serialize results
        results = []
        for submission in submissions:
            results.append({
                'id': submission.id,
                'name': submission.name if not submission.anonymized else 'ANONYMIZED',
                'email': submission.email if not submission.anonymized else 'ANONYMIZED',
                'phone': submission.phone if not submission.anonymized else 'ANONYMIZED',
                'country': submission.country,
                'submitted_at': submission.submitted_at.isoformat(),
                'short_summary': submission.short_summary,
                'anonymized': submission.anonymized,
            })
        
        # Log data access
        log_security_event(
            'DATA_ACCESS', 'MEDIUM',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, f'Admin accessed {len(results)} submissions',
            {
                'record_count': len(results),
                'total_available': total_count,
                'filters_applied': len(validated_filters),
                'admin_user': request.user.username
            }
        )
        
        return Response({
            'success': True,
            'results': results,
            'count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': (total_count + page_size - 1) // page_size,
        })
        
    except Exception as e:
        logger.error(f"Admin submissions error: {e}")
        log_security_event(
            'API_ERROR', 'HIGH',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, 'Admin submissions failed due to system error'
        )
        return Response({
            'success': False,
            'message': 'Failed to load submissions'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAdminUser])
@csrf_exempt
@never_cache
def submission_detail(request, pk):
    """Get detailed submission information"""
    
    try:
        # Validate and sanitize PK
        try:
            submission_id = int(pk)
            if submission_id <= 0:
                raise ValueError("Invalid ID")
        except (ValueError, TypeError):
            log_security_event(
                'SUSPICIOUS_ACTIVITY', 'HIGH',
                get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                request.user, f'Invalid submission ID in detail request: {pk}'
            )
            return Response({
                'success': False,
                'message': 'Invalid submission ID'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # SQL injection check
        is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(str(pk))
        if is_injection:
            log_security_event(
                'SQL_INJECTION_ATTEMPT', 'CRITICAL',
                get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                request.user, f'SQL injection in detail PK: {pattern}'
            )
            return Response({
                'success': False,
                'message': 'Security violation detected'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get submission
        submission = Submission.objects.get(pk=submission_id)
        
        # Log access
        log_security_event(
            'SENSITIVE_DATA_ACCESS', 'HIGH',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, 'Admin viewed submission details',
            {
                'submission_id': submission_id,
                'submission_uuid': str(submission.uuid),
                'admin_user': request.user.username
            }
        )
        
        # Return detailed data
        return Response({
            'success': True,
            'id': submission.id,
            'uuid': str(submission.uuid),
            'step1': submission.step1,
            'step2': submission.step2,
            'step3': submission.step3,
            'step4': submission.step4,
            'step5': submission.step5,
            'step6': submission.step6,
            'step7': submission.step7,
            'step8': submission.step8,
            'name': submission.name,
            'email': submission.email,
            'phone': submission.phone,
            'country': submission.country,
            'submitted_at': submission.submitted_at.isoformat(),
            'anonymized': submission.anonymized,
            'data_classification': submission.data_classification,
        })
        
    except Submission.DoesNotExist:
        log_security_event(
            'SUSPICIOUS_ACTIVITY', 'MEDIUM',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, f'Admin attempted to access non-existent submission: {pk}'
        )
        return Response({
            'success': False,
            'message': 'Submission not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Submission detail error: {e}")
        return Response({
            'success': False,
            'message': 'Unable to retrieve submission details'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
@permission_classes([IsAdminUser])
@csrf_exempt
@never_cache
def delete_submission(request, pk):
    """Delete a single submission"""
    
    try:
        # Validate PK
        try:
            submission_id = int(pk)
            if submission_id <= 0:
                raise ValueError("Invalid ID")
        except (ValueError, TypeError):
            return Response({
                'success': False,
                'message': 'Invalid submission ID'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # SQL injection check
        is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(str(pk))
        if is_injection:
            log_security_event(
                'SQL_INJECTION_ATTEMPT', 'CRITICAL',
                get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                request.user, f'SQL injection in delete PK: {pattern}'
            )
            return Response({
                'success': False,
                'message': 'Security violation detected'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Get and delete submission
        submission = Submission.objects.get(pk=submission_id)
        
        # Log before deletion
        log_security_event(
            'DATA_DELETION', 'CRITICAL',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, 'Admin deleted submission',
            {
                'submission_id': submission_id,
                'submission_uuid': str(submission.uuid),
                'admin_user': request.user.username,
                'submitted_at': submission.submitted_at.isoformat(),
            }
        )
        
        submission.delete()
        
        return Response({
            'success': True,
            'message': 'Submission deleted successfully'
        })
        
    except Submission.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Submission not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Delete error: {e}")
        return Response({
            'success': False,
            'message': 'Deletion failed'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAdminUser])
@csrf_exempt
@never_cache
def download_submissions_excel(request):
    """Download submissions as Excel file"""
    
    try:
        # SQL injection detection on query parameters
        for param_name, param_value in request.GET.items():
            if param_value:
                is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(str(param_value))
                if is_injection:
                    log_security_event(
                        'SQL_INJECTION_ATTEMPT', 'CRITICAL',
                        get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                        request.user, f'SQL injection in download parameter {param_name}: {pattern}'
                    )
                    return Response({
                        'success': False,
                        'message': 'Security violation detected in download parameters'
                    }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate filter parameters
        ALLOWED_FILTERS = {
            'date_from': r'^\d{4}-\d{2}-\d{2}$',
            'date_to': r'^\d{4}-\d{2}-\d{2}$',
            'country': r'^[A-Z]{2}$',
            'service_type': r'^[a-zA-Z\s\-]{1,50}$',
        }
        
        validated_filters = {}
        for filter_key, pattern in ALLOWED_FILTERS.items():
            if filter_key in request.GET:
                filter_value = str(request.GET[filter_key]).strip()
                if re.match(pattern, filter_value):
                    validated_filters[filter_key] = filter_value
        
        # Build queryset
        queryset = Submission.objects.all()
        
        # Apply filters
        if validated_filters.get('date_from'):
            try:
                start_date = timezone.datetime.strptime(validated_filters['date_from'], '%Y-%m-%d').date()
                queryset = queryset.filter(submitted_at__date__gte=start_date)
            except ValueError:
                pass
        
        if validated_filters.get('date_to'):
            try:
                end_date = timezone.datetime.strptime(validated_filters['date_to'], '%Y-%m-%d').date()
                queryset = queryset.filter(submitted_at__date__lte=end_date)
            except ValueError:
                pass
        
        if validated_filters.get('country'):
            queryset = queryset.filter(country=validated_filters['country'])
        
        # Limit download size
        MAX_DOWNLOAD_RECORDS = 10000
        if queryset.count() > MAX_DOWNLOAD_RECORDS:
            return Response({
                'success': False,
                'message': f'Too many records to download. Limit: {MAX_DOWNLOAD_RECORDS}. Please apply filters to reduce the dataset.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get submissions
        submissions = queryset.order_by('-submitted_at')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Form Submissions"
        
        # Headers
        headers = [
            'ID', 'UUID', 'Company Name', 'Service Type', 'Issue Timeline',
            'Company Acknowledgment', 'Primary Goal', 'How Heard About Us',
            'Communication Method', 'Case Summary', 'Name', 'Email', 'Phone',
            'Country', 'Submitted At', 'Anonymized', 'Data Classification'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row, submission in enumerate(submissions, 2):
            safe_data = [
                str(submission.id),
                str(submission.uuid),
                escape(str(submission.step1 or ''))[:500],
                escape(str(submission.step2 or ''))[:100],
                escape(str(submission.step3 or ''))[:100],
                escape(str(submission.step4 or ''))[:100],
                escape(str(submission.step5 or ''))[:100],
                escape(str(submission.step6 or ''))[:100],
                escape(str(submission.step7 or ''))[:100],
                escape(str(submission.step8 or ''))[:1000],
                escape(str(submission.name or ''))[:200],
                escape(str(submission.email or ''))[:254],
                escape(str(submission.phone or ''))[:50],
                str(submission.country or ''),
                submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Yes' if submission.anonymized else 'No',
                str(getattr(submission, 'data_classification', 'CONFIDENTIAL'))
            ]
            
            for col, value in enumerate(safe_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = min(20, max(10, len(headers[col-1]) + 2))
        
        # Add metadata
        ws.cell(row=submissions.count() + 3, column=1, 
                value=f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} UTC")
        ws.cell(row=submissions.count() + 4, column=1, 
                value=f"User: {request.user.username}")
        ws.cell(row=submissions.count() + 5, column=1, 
                value="CONFIDENTIAL - Authorized Personnel Only")
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Log download
        record_count = submissions.count()
        log_security_event(
            'DATA_EXPORT', 'HIGH',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, f'Admin downloaded {record_count} submissions as Excel',
            {
                'admin_user': request.user.username,
                'record_count': record_count,
                'filters_applied': len(validated_filters),
                'file_size_bytes': len(buffer.getvalue())
            }
        )
        
        # Generate filename
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = f"submissions_{timestamp}.xlsx"
        
        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_filename}"'
        response['X-Content-Type-Options'] = 'nosniff'
        response['X-Frame-Options'] = 'DENY'
        
        return response
        
    except Exception as e:
        logger.error(f"Excel download error: {e}")
        log_security_event(
            'API_ERROR', 'HIGH',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, 'Excel download failed due to system error'
        )
        return Response({
            'success': False,
            'message': 'Failed to generate Excel file. Please try again or contact administrator.'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAdminUser])
@csrf_exempt
@never_cache
def delete_all_submissions(request):
    """Bulk delete all submissions"""
    
    try:
        # SQL injection detection on request body
        request_body = request.body.decode('utf-8') if request.body else ''
        is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(request_body)
        
        if is_injection:
            log_security_event(
                'SQL_INJECTION_ATTEMPT', 'CRITICAL',
                get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                request.user, f'SQL injection in bulk delete request: {pattern}'
            )
            return Response({
                'success': False,
                'message': 'Security violation detected'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate confirmation
        confirmation = request.data.get('confirmation')
        if not confirmation:
            return Response({
                'success': False,
                'message': 'Confirmation token required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # SQL injection check on confirmation
        is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(str(confirmation))
        if is_injection:
            log_security_event(
                'SQL_INJECTION_ATTEMPT', 'CRITICAL',
                get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                request.user, f'SQL injection in confirmation token: {pattern}'
            )
            return Response({
                'success': False,
                'message': 'Invalid confirmation token format'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Exact string match
        REQUIRED_CONFIRMATION = 'delete_permanently'
        if str(confirmation).strip() != REQUIRED_CONFIRMATION:
            log_security_event(
                'SUSPICIOUS_ACTIVITY', 'HIGH',
                get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                request.user, f'Invalid bulk delete confirmation: {confirmation[:20]}...'
            )
            return Response({
                'success': False,
                'message': 'Invalid confirmation code. Please type exactly: delete_permanently'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Rate limiting
        client_ip = get_client_ip(request)
        safe_ip = re.sub(r'[^0-9\.\:]', '_', client_ip)
        if SQLSecureRateLimiter.is_rate_limited(f"bulk_delete_{safe_ip}", 1, 86400):
            log_security_event(
                'RATE_LIMIT', 'HIGH',
                get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                request.user, 'Bulk delete rate limit exceeded'
            )
            return Response({
                'success': False,
                'message': 'Bulk deletion rate limit exceeded. Only 1 bulk deletion allowed per day.'
            }, status=status.HTTP_429_TOO_MANY_REQUESTS)
        
        # Count records before deletion
        total_count = Submission.objects.count()
        
        if total_count == 0:
            return Response({
                'success': True,
                'message': 'No submissions to delete'
            })
        
        # Log before deletion
        log_security_event(
            'BULK_DATA_DELETION', 'CRITICAL',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, f'Admin initiated bulk deletion of {total_count} submissions',
            {
                'admin_user': request.user.username,
                'submission_count': total_count,
                'confirmation_provided': REQUIRED_CONFIRMATION,
                'deletion_timestamp': timezone.now().isoformat(),
                'client_ip': client_ip
            }
        )
        
        # Delete using atomic transaction
        with transaction.atomic():
            sample_submissions = list(Submission.objects.values(
                'uuid', 'submitted_at', 'country', 'anonymized'
            )[:10])
            
            deleted_count, _ = Submission.objects.all().delete()
            
            log_security_event(
                'BULK_DATA_DELETION', 'CRITICAL',
                get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                request.user, f'Successfully deleted {deleted_count} submissions',
                {
                    'admin_user': request.user.username,
                    'deleted_count': deleted_count,
                    'sample_data': sample_submissions,
                    'completion_timestamp': timezone.now().isoformat()
                }
            )
        
        return Response({
            'success': True,
            'message': f'Successfully deleted {deleted_count} submissions',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        logger.error(f"Bulk delete error: {e}")
        log_security_event(
            'API_ERROR', 'HIGH',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, f'Bulk deletion failed due to system error: {str(e)[:100]}'
        )
        return Response({
            'success': False,
            'message': 'Bulk deletion failed due to system error'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAdminUser])
@csrf_exempt
@never_cache
def get_filter_options(request):
    """Get filter options for admin interface"""
    
    try:
        submissions = Submission.objects.all()
        
        # Extract unique service types
        service_types = set()
        for submission in submissions.only('step2'):
            if submission.step2:
                service_type = str(submission.step2).strip()
                service_type = escape(service_type)
                if len(service_type) <= 100 and re.match(r'^[a-zA-Z0-9\s\-/]+$', service_type):
                    service_types.add(service_type)
        
        # Extract unique issue timeframes
        issue_timeframes = set()
        for submission in submissions.only('step3'):
            if submission.step3:
                timeframe = str(submission.step3).strip()
                timeframe = escape(timeframe)
                if len(timeframe) <= 100 and re.match(r'^[a-zA-Z0-9\s\-/]+$', timeframe):
                    issue_timeframes.add(timeframe)
        
        # Extract countries
        countries = []
        for country_code in submissions.values_list('country', flat=True).distinct():
            if country_code:
                safe_country = str(country_code).strip().upper()
                if re.match(r'^[A-Z]{2}$', safe_country):
                    countries.append({
                        'code': safe_country,
                        'display': safe_country
                    })
        
        # Static options
        acknowledgments = ['Yes, full-time', 'Yes, part-time', 'No', 'Partially']
        primary_goals = ['Financial Recovery', 'Legal Action', 'Information', 'Compensation', 'Other']
        heard_abouts = ['Social Media', 'Friend Referral', 'Online Search', 'Advertisement', 'Legal Referral', 'Other']
        communication_methods = ['Email', 'Phone', 'Text Message', 'Video Call', 'Secure Portal']
        
        # Log access
        log_security_event(
            'DATA_ACCESS', 'LOW',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, 'Admin accessed filter options',
            {
                'admin_user': request.user.username,
                'service_types_count': len(service_types),
                'timeframes_count': len(issue_timeframes),
                'countries_count': len(countries)
            }
        )
        
        return Response({
            'success': True,
            'service_types': sorted(list(service_types)),
            'issue_timeframes': sorted(list(issue_timeframes)),
            'acknowledgments': acknowledgments,
            'primary_goals': primary_goals,
            'heard_abouts': heard_abouts,
            'communication_methods': communication_methods,
            'countries': sorted(countries, key=lambda x: x['code'])
        })
        
    except Exception as e:
        logger.error(f"Filter options error: {e}")
        log_security_event(
            'API_ERROR', 'MEDIUM',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, 'Filter options failed due to system error'
        )
        return Response({
            'success': False,
            'message': 'Failed to load filter options'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAdminUser])
@csrf_exempt
@never_cache
def submission_stats(request):
    """Get submission statistics"""
    
    try:
        # SQL injection detection on parameters
        for param_name, param_value in request.GET.items():
            if param_value:
                is_injection, pattern, severity = SQLInjectionDetector.detect_sql_injection(str(param_value))
                if is_injection:
                    log_security_event(
                        'SQL_INJECTION_ATTEMPT', 'CRITICAL',
                        get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
                        request.user, f'SQL injection in stats parameter {param_name}: {pattern}'
                    )
                    return Response({
                        'success': False,
                        'message': 'Security violation detected in statistics parameters'
                    }, status=status.HTTP_403_FORBIDDEN)
        
        # Validate filter parameters
        ALLOWED_FILTERS = {
            'date_from': r'^\d{4}-\d{2}-\d{2}$',
            'date_to': r'^\d{4}-\d{2}-\d{2}$',
            'date_preset': r'^(today|1_week|2_weeks|30_days)$',
            'country': r'^[A-Z]{2}$',
        }
        
        validated_filters = {}
        for filter_key, pattern in ALLOWED_FILTERS.items():
            if filter_key in request.GET:
                filter_value = str(request.GET[filter_key]).strip()
                if re.match(pattern, filter_value):
                    validated_filters[filter_key] = filter_value
        
        # Base queryset
        queryset = Submission.objects.all()
        
        # Apply date filtering
        now = timezone.now()
        date_range = {'from': '', 'to': '', 'preset': ''}
        
        if validated_filters.get('date_preset'):
            preset = validated_filters['date_preset']
            date_range['preset'] = preset
            
            if preset == 'today':
                queryset = queryset.filter(submitted_at__date=now.date())
            elif preset == '1_week':
                week_ago = now - timedelta(days=7)
                queryset = queryset.filter(submitted_at__gte=week_ago)
            elif preset == '2_weeks':
                weeks_ago = now - timedelta(days=14)
                queryset = queryset.filter(submitted_at__gte=weeks_ago)
            elif preset == '30_days':
                month_ago = now - timedelta(days=30)
                queryset = queryset.filter(submitted_at__gte=month_ago)
        
        elif validated_filters.get('date_from') and validated_filters.get('date_to'):
            try:
                start_date = timezone.datetime.strptime(validated_filters['date_from'], '%Y-%m-%d').date()
                end_date = timezone.datetime.strptime(validated_filters['date_to'], '%Y-%m-%d').date()
                queryset = queryset.filter(submitted_at__date__range=[start_date, end_date])
                date_range['from'] = validated_filters['date_from']
                date_range['to'] = validated_filters['date_to']
            except ValueError:
                pass
        
        if validated_filters.get('country'):
            queryset = queryset.filter(country=validated_filters['country'])
        
        # Calculate statistics
        total_submissions = queryset.count()
        
        # Service type breakdown
        service_type_breakdown = {}
        for submission in queryset.only('step2'):
            if submission.step2:
                service_type = escape(str(submission.step2).strip())
                if len(service_type) <= 100:
                    service_type_breakdown[service_type] = service_type_breakdown.get(service_type, 0) + 1
        
        # Country breakdown
        country_breakdown = {}
        country_counts = queryset.values('country').annotate(count=Count('country'))
        for item in country_counts:
            if item['country']:
                safe_country = str(item['country']).strip().upper()
                if re.match(r'^[A-Z]{2}$', safe_country):
                    country_breakdown[safe_country] = item['count']
        
        # Issue timeframe breakdown
        issue_timeframe_breakdown = {}
        for submission in queryset.only('step3'):
            if submission.step3:
                timeframe = escape(str(submission.step3).strip())
                if len(timeframe) <= 100:
                    issue_timeframe_breakdown[timeframe] = issue_timeframe_breakdown.get(timeframe, 0) + 1
        
        # Daily submissions (last 7 days)
        daily_submissions = []
        for i in range(7):
            date = (now - timedelta(days=i)).date()
            count = queryset.filter(submitted_at__date=date).count()
            daily_submissions.append({
                'date': date.isoformat(),
                'count': count
            })
        daily_submissions.reverse()  # Chronological order
        
        stats_data = {
            'total_submissions': total_submissions,
            'service_type_breakdown': dict(sorted(service_type_breakdown.items())),
            'country_breakdown': dict(sorted(country_breakdown.items())),
            'issue_timeframe_breakdown': dict(sorted(issue_timeframe_breakdown.items())),
            'daily_submissions': daily_submissions,
            'date_range': date_range
        }
        
        # Log access
        log_security_event(
            'DATA_ACCESS', 'MEDIUM',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, f'Admin accessed statistics for {total_submissions} submissions',
            {
                'admin_user': request.user.username,
                'total_submissions': total_submissions,
                'filters_applied': len(validated_filters),
                'date_range': date_range['preset'] or 'custom'
            }
        )
        
        return Response({
            'success': True,
            **stats_data
        })
        
    except Exception as e:
        logger.error(f"Statistics error: {e}")
        log_security_event(
            'API_ERROR', 'MEDIUM',
            get_client_ip(request), request.META.get('HTTP_USER_AGENT', ''),
            request.user, 'Statistics failed due to system error'
        )
        return Response({
            'success': False,
            'message': 'Failed to load statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)